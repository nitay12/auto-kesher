import openpyxl
from openpyxl import load_workbook as lwb
import http.client
import urllib.parse
import json
import googlemaps
import re

from sensitive import gmapsKey, gmailPassword, sender_address

import smtplib, email
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

def addresses_sort(file_path):
    #Constants
    N_MIN_LAT = 32.100766
    E_MIN_LAT_A = 32.064696
    E_MAX_LAT_A = 32.079732
    E_MIN_LON_A = 34.793462
    E_MAX_LAT_B = 32.064696
    E_MIN_LON_B = 34.785705
    S_MAX_LAT = 32.069
    S_MAX_LON = 34.786273
    MAP_URL = "http://maps.google.com/maps?"
    G_MAPS = googlemaps.Client(key=gmapsKey)
    #Getting columns from JSON
    with open ("config.json") as config:
            settings = json.load(config)
            columns = settings["columns"]
            date_col = int(columns["dateDead"])
            family_col = int(columns["familyName"])
            name_col = int(columns["firstName"])
            street_col = int(columns["street"])
            street_num_col = int(columns["street_num"])
            city_col = int(columns["city"])
            conection_col = int(columns["deadConection"])
            phone_col = int(columns["phone"])
    #conn = http.client.HTTPConnection('api.positionstack.com')
    north_addresses = ""
    west_addresses = ""
    south_addresses = ""
    east_addresses = ""
    errors = []

    north_url = [MAP_URL]
    east_url = [MAP_URL]
    west_url = [MAP_URL]
    south_url = [MAP_URL]

    def URL_Encoded_add(employee_url ,address):
        encodedAddress = re.sub('\W+',' ', address)
        encodedAddress = encodedAddress.replace(" ", "+")
        if "saddr" not in employee_url[0]:
            print ("adding saddr: ")
            employee_url[0]+="saddr="+encodedAddress
        elif "daddr" not in employee_url[0]:
            print ("adding daddr: ")
            employee_url[0]+= "&daddr=" + encodedAddress
        elif "daddr" in employee_url[0]:
            print ("adding to: ")
            employee_url[0]+= "+to:" + encodedAddress
        print(encodedAddress)
        print(employee_url[0])

    # getting the adresses from Excel file
    wb = lwb(file_path)
    ws = wb.active
    max_row = ws.max_row
    for i in range(1, max_row+1):  # TODO: Search for the Death_street&Street_number columns
        dateDead = ws.cell(row=i, column=date_col)
        familyName = ws.cell(row=i, column=family_col)
        firstName = ws.cell(row=i, column=name_col)
        street = ws.cell(row=i, column=street_col)
        street_num = ws.cell(row=i, column=street_num_col)
        city = ws.cell(row=i, column=city_col)
        deadConection = ws.cell(row=i, column=conection_col)
        phone = ws.cell(row=i, column=phone_col)
        adressString = str(street.value)+" "+str(street_num.value)+" "+str(city.value)
        mail =  "כתובת: {} \n שם: {} שם משפחה:{} \n תאריך: {} \n טלפון: {} קרבה: {} \n\n".format(adressString, firstName.value, familyName.value, dateDead.value, phone.value, deadConection.value)
        print(mail)

        if adressString == "None None None":
            continue
        else:

    # Forward geocoding the adress
            try:  
                res = G_MAPS.geocode(adressString)  # conn.getresponse()
                lat = float(res[0]['geometry']['location']['lat'])
                print(lat)
                lon = float(res[0]['geometry']['location']['lng'])
                print(lon)

            #Sort the adresses by workers
                #Sorting addresses for North Worker
                if lat > N_MIN_LAT: 
                    north_addresses+=mail
                    URL_Encoded_add(north_url,adressString)
                    print(adressString,"added to URL")
                    print(adressString+" added to north")
                #Sorting addresses for East Worker
                elif (E_MIN_LAT_A < lat <= E_MAX_LAT_A and lon >= E_MIN_LON_A) or (lat < E_MAX_LAT_B and lon > E_MIN_LON_B):
                    east_addresses+=mail
                    URL_Encoded_add(east_url,adressString)
                    print(adressString+" added to east")
                #Sorting addresses foB South Worker
                elif lat < S_MAX_LAT  and lon < S_MAX_LON:
                    URL_Encoded_add(south_url,adressString)
                    south_addresses+=mail
                    print(adressString+" added to south")
                #Sorting addresses for West Worker
                else:
                    west_addresses+=mail
                    URL_Encoded_add(west_url,adressString)
                    print(adressString+" added to west")
            except:
                errors.append(adressString)
    north_addresses+= "קישור: "+north_url[0]+"\n"
    east_addresses+= "קישור: "+east_url[0]+"\n"
    west_addresses+= "קישור: "+west_url[0]+"\n"
    south_addresses+="קישור: "+south_url[0]+"\n"
    return {"north":north_addresses,"east":east_addresses,"south":south_addresses,"west":west_addresses}

def send_mail(workerEmail, content, zone):
    #The mail addresses and password
    receiver_address = workerEmail
    #Setup the MIME
    message = MIMEMultipart()
    message['From'] = sender_address
    message['To'] = receiver_address
    message['Subject'] =  zone   #The subject line
    #The body and the attachments for the mail
    mail_content = content
    message.attach(MIMEText(mail_content, 'plain'))
    #Create SMTP session for sending the mail
    session = smtplib.SMTP('smtp.gmail.com', 587) #use gmail with port
    session.starttls() #enable security
    session.login(sender_address, gmailPassword) #login with mail_id and password
    text = message.as_string()
    session.sendmail(sender_address, receiver_address, text)
    session.quit()